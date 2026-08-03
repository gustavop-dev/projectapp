"""API tests for the accounting CSV/XLSX export endpoints."""
from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from content.models import CardBalanceSnapshot, IncomeRecord
from content.views.accounting_export import XLSX_CONTENT_TYPE


def csv_lines(response):
    body = response.content.decode('utf-8')
    assert body.startswith('\ufeff')
    return [line for line in body.lstrip('\ufeff').splitlines() if line]


@pytest.mark.django_db
class TestExportRecords:
    def test_requires_superuser(self, admin_client):
        response = admin_client.get('/api/accounting/export/?section=income')
        assert response.status_code == 403

    def test_invalid_section_returns_400(self, super_client):
        response = super_client.get('/api/accounting/export/?section=nope')
        assert response.status_code == 400
        assert response.data['code'] == 'invalid_section'

    def test_invalid_format_returns_400(self, super_client):
        response = super_client.get(
            '/api/accounting/export/?section=income&file_format=pdf',
        )
        assert response.status_code == 400
        assert response.data['code'] == 'invalid_format'

    def test_csv_headers_in_spanish_and_bom(self, super_client, make_income):
        make_income(concept='Kore - Inicio 40%')
        response = super_client.get('/api/accounting/export/?section=income')
        assert response.status_code == 200
        assert response['Content-Type'].startswith('text/csv')
        assert 'contabilidad_income_' in response['Content-Disposition']
        lines = csv_lines(response)
        assert lines[0].split(',')[:5] == [
            'Concepto', 'Tipo', 'Estado de cobro', 'Contabilidad', 'Mes',
        ]
        assert 'Kore - Inicio 40%' in lines[1]

    def test_csv_respects_filters(self, super_client, make_income):
        make_income(
            concept='Company liquid', kind=IncomeRecord.Kind.LIQUID,
        )
        make_income(
            concept='Personal Tavo', kind=IncomeRecord.Kind.LIQUID,
            ledger=IncomeRecord.Ledger.GUSTAVO,
            total_amount=Decimal('100.00'),
            gustavo_amount=Decimal('100.00'),
            carlos_amount=Decimal('0.00'),
        )
        response = super_client.get(
            '/api/accounting/export/?section=income&kind=liquid&ledger=gustavo',
        )
        lines = csv_lines(response)
        assert len(lines) == 2  # header + 1 row
        assert 'Personal Tavo' in lines[1]

    def test_choice_filter_accepts_comma_multi(self, super_client, make_income):
        make_income(concept='A', kind=IncomeRecord.Kind.LIQUID)
        make_income(
            concept='B', kind=IncomeRecord.Kind.LIQUID,
            ledger=IncomeRecord.Ledger.CARLOS,
            total_amount=Decimal('10.00'),
            gustavo_amount=Decimal('0.00'),
            carlos_amount=Decimal('10.00'),
        )
        make_income(
            concept='C', kind=IncomeRecord.Kind.LIQUID,
            ledger=IncomeRecord.Ledger.GUSTAVO,
            total_amount=Decimal('10.00'),
            gustavo_amount=Decimal('10.00'),
            carlos_amount=Decimal('0.00'),
        )
        response = super_client.get(
            '/api/accounting/export/?section=income&ledger=gustavo,carlos',
        )
        lines = csv_lines(response)
        assert len(lines) == 3
        assert all('A,' not in line for line in lines[1:])

    def test_empty_export_has_only_header(self, super_client):
        response = super_client.get('/api/accounting/export/?section=expense')
        lines = csv_lines(response)
        assert len(lines) == 1

    def test_csv_income_carries_collection_state_labels(
        self, super_client, make_income,
    ):
        paid = make_income(concept='Esperado pagado')
        make_income(
            concept='Pago total', kind=IncomeRecord.Kind.LIQUID,
            expected_income=paid,
        )
        make_income(concept='Esperado sin pagos')
        make_income(concept='Liquido suelto', kind=IncomeRecord.Kind.LIQUID)
        response = super_client.get('/api/accounting/export/?section=income')
        lines = csv_lines(response)
        state_index = lines[0].split(',').index('Estado de cobro')
        states = {
            line.split(',')[0]: line.split(',')[state_index]
            for line in lines[1:]
        }
        assert states['Esperado pagado'] == 'Pagado'
        assert states['Esperado sin pagos'] == 'Pendiente'
        assert states['Liquido suelto'] == ''
        assert states['Pago total'] == ''

    def test_csv_income_partial_collection_state(
        self, super_client, make_income,
    ):
        partial = make_income(concept='Esperado parcial')
        make_income(
            concept='Abono', kind=IncomeRecord.Kind.LIQUID,
            expected_income=partial,
            total_amount=Decimal('400000.00'),
            gustavo_amount=Decimal('200000.00'),
            carlos_amount=Decimal('200000.00'),
        )
        response = super_client.get('/api/accounting/export/?section=income')
        lines = csv_lines(response)
        state_index = lines[0].split(',').index('Estado de cobro')
        row = next(l for l in lines[1:] if l.startswith('Esperado parcial'))
        assert row.split(',')[state_index] == 'Parcial'

    def test_xlsx_export_opens_and_has_rows(self, super_client, make_income):
        make_income(concept='Ingreso XLSX')
        response = super_client.get(
            '/api/accounting/export/?section=income&file_format=xlsx',
        )
        assert response.status_code == 200
        assert response['Content-Disposition'].endswith('.xlsx"')
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook['Ingresos']
        assert sheet.cell(row=1, column=1).value == 'Concepto'
        assert sheet.cell(row=2, column=1).value == 'Ingreso XLSX'


@pytest.mark.django_db
class TestExportWorkbook:
    def test_requires_superuser(self, admin_client):
        response = admin_client.get('/api/accounting/export/workbook/')
        assert response.status_code == 403

    def test_invalid_year_returns_400(self, super_client):
        """Catches the error code silently drifting (accounting_export.py:76-84)
        — clients that branch on `code` would break even though the status
        stayed 400.
        """
        response = super_client.get('/api/accounting/export/workbook/?year=x')
        assert response.status_code == 400
        assert response.data['code'] == 'invalid_year'

    def test_workbook_ingresos_and_resumen_sheets_carry_real_data(
        self, super_client, make_income,
    ):
        """Catches a section rendering only headers with no row data, a wrong
        Content-Type header, or the summary block silently dropping/renaming
        the 'Ingresos esperados' row (accounting_export_service.py:230-249) —
        none of which the sheetnames-only assertions below would catch.
        """
        make_income(
            kind=IncomeRecord.Kind.LIQUID, concept='Kore v2 anticipo distintivo',
        )
        response = super_client.get(
            '/api/accounting/export/workbook/?year=2026',
        )
        assert response['Content-Type'] == XLSX_CONTENT_TYPE

        workbook = load_workbook(BytesIO(response.content))
        ingresos_values = [
            cell.value for row in workbook['Ingresos'].iter_rows(min_row=2)
            for cell in row
        ]
        assert 'Kore v2 anticipo distintivo' in ingresos_values

        resumen_labels = [
            cell.value for row in workbook['Resumen'].iter_rows()
            for cell in row
        ]
        assert 'Ingresos esperados' in resumen_labels

    def test_workbook_income_sheet_carries_collection_state(
        self, super_client, make_income,
    ):
        """Catches the year workbook regressing to the raw manager: without
        base_queryset the paid_amount annotation is missing and this column
        would crash (accounting_export.py:86-94).
        """
        settled = make_income(concept='Esperado workbook')
        make_income(
            concept='Pago workbook', kind=IncomeRecord.Kind.LIQUID,
            expected_income=settled,
        )
        response = super_client.get('/api/accounting/export/workbook/?year=2026')
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook['Ingresos']
        headers = [cell.value for cell in sheet[1]]
        state_column = headers.index('Estado de cobro') + 1
        states = {
            sheet.cell(row=i, column=1).value:
                sheet.cell(row=i, column=state_column).value
            for i in range(2, sheet.max_row + 1)
        }
        assert states['Esperado workbook'] == 'Pagado'

    def test_workbook_contains_summary_and_section_sheets(
        self, super_client, make_income, make_expense,
    ):
        make_income(kind=IncomeRecord.Kind.LIQUID, concept='Ingreso 2026')
        make_expense(concept='Gasto 2026')
        CardBalanceSnapshot.objects.create(
            snapshot_date=date(2026, 6, 17), card_name='T.C 0064',
            available_amount=Decimal('100.00'), debt_amount=Decimal('900.00'),
        )
        response = super_client.get(
            '/api/accounting/export/workbook/?year=2026',
        )
        assert response.status_code == 200
        workbook = load_workbook(BytesIO(response.content))
        assert workbook.sheetnames == [
            'Resumen', 'Ingresos', 'Gastos', 'Hostings', 'Bolsillo',
            'Recurrentes', 'Ads', 'Tarjetas', 'Extractos TC',
            'Transacciones TC',
        ]
        assert workbook['Resumen'].cell(row=1, column=1).value == 'Resumen 2026'
        assert workbook['Tarjetas'].cell(row=2, column=1).value == 'T.C 0064'

    def test_workbook_scopes_sections_to_year(self, super_client, make_income):
        make_income(
            kind=IncomeRecord.Kind.LIQUID, concept='Viejo',
            period_date=date(2025, 5, 1),
        )
        response = super_client.get(
            '/api/accounting/export/workbook/?year=2026',
        )
        workbook = load_workbook(BytesIO(response.content))
        assert workbook['Ingresos'].max_row == 1


@pytest.mark.django_db
class TestStatementExportSections:
    def test_csv_export_of_statements(self, super_client):
        from content.models import CreditCardStatement

        CreditCardStatement.objects.create(
            card_name='Visa Bancolombia', period_date=date(2026, 6, 1),
            purchases_total=Decimal('100000.00'),
        )
        response = super_client.get(
            '/api/accounting/export/?section=statement&file_format=csv',
        )
        assert response.status_code == 200
        body = response.content.decode('utf-8-sig')
        assert 'Visa Bancolombia' in body
        assert 'Total compras' in body

    def test_workbook_includes_statement_transactions_sheet(self, super_client):
        from content.models import CreditCardStatement, CreditCardTransaction

        statement = CreditCardStatement.objects.create(
            card_name='Visa Bancolombia', period_date=date(2026, 6, 1),
            purchases_total=Decimal('100000.00'),
        )
        CreditCardTransaction.objects.create(
            statement=statement, transaction_date=date(2026, 6, 5),
            raw_description='PAYU*NETFLIX', merchant_name='Netflix',
            amount=Decimal('44900.00'),
        )
        response = super_client.get(
            '/api/accounting/export/workbook/?year=2026',
        )
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook['Transacciones TC']
        assert sheet.cell(row=2, column=1).value == 'Visa Bancolombia'
        assert sheet.cell(row=2, column=5).value == 'Netflix'


@pytest.mark.django_db
class TestExpenseDeductionExport:
    def test_csv_carries_deduction_type_and_origin_income(
        self, super_client, make_income, make_expense,
    ):
        income = make_income(concept='Hosting Jimmy Junio')
        make_expense(concept='Hosting mensual')
        make_expense(
            concept='Comisión Wompi',
            deduction_type='gateway_fee',
            source_income=income,
            total_amount=Decimal('4854.00'),
            gustavo_amount=Decimal('2427.00'),
            carlos_amount=Decimal('2427.00'),
        )

        response = super_client.get('/api/accounting/export/?section=expense')

        lines = csv_lines(response)
        headers = lines[0].split(',')
        type_index = headers.index('Tipo de deducción')
        origin_index = headers.index('Ingreso origen')
        rows = {line.split(',')[0]: line.split(',') for line in lines[1:]}
        assert rows['Comisión Wompi'][type_index] == 'Comisión plataforma de pago'
        assert rows['Comisión Wompi'][origin_index] == 'Hosting Jimmy Junio'
        assert rows['Hosting mensual'][type_index] == ''
        assert rows['Hosting mensual'][origin_index] == ''

    def test_csv_respects_the_deduction_type_filter(
        self, super_client, make_expense,
    ):
        make_expense(concept='Hosting mensual')
        make_expense(
            concept='Retención DIAN',
            deduction_type='withholding',
            total_amount=Decimal('70000.00'),
            gustavo_amount=Decimal('35000.00'),
            carlos_amount=Decimal('35000.00'),
        )

        response = super_client.get(
            '/api/accounting/export/?section=expense&deduction_type=withholding',
        )

        lines = csv_lines(response)
        assert len(lines) == 2
        assert lines[1].startswith('Retención DIAN')
