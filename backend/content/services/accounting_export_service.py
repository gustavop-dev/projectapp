"""CSV/XLSX export for the accounting module.

Sections mirror the `_ENTITIES` keys of content.views.accounting; each
one declares its Spanish headers and how to read every cell from a
record (field name or callable). Search parity note: server-side `q`
uses icontains over the section's search_fields, matching the panel's
client-side substring filter closely enough for exports.
"""
import csv
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from content.models import IncomeRecord
from content.serializers.accounting import (
    PAYMENT_STATUS_LABELS,
    payment_status_for,
)
from content.services.accounting_service import dashboard_summary
from content.utils import today_bogota

MONEY_FORMAT = '#,##0'
HEADER_FONT = Font(bold=True)


def _income_payment_status(record):
    """Estado de cobro cell; hard access to the paid_amount annotation on
    purpose — export querysets must come from base_queryset (its contract)."""
    if record.kind != IncomeRecord.Kind.EXPECTED:
        return ''
    return PAYMENT_STATUS_LABELS.get(
        payment_status_for(record.paid_amount, record.total_amount), '',
    )


def _hosting_client(record):
    """Linked client's display name; blank while the hosting is pending."""
    if not record.client_id:
        return ''
    from accounts.services.proposal_client_service import (
        build_client_display_name,
    )

    return build_client_display_name(record.client)


def _record_project(record):
    """Linked project's name; blank when the row has none (a cobro por
    diagnóstico legitimately does)."""
    return record.project.name if record.project_id else ''


_POCKET_ATTRIBUTION_LABELS = {
    'company': 'Empresa',
    'gustavo': 'Gustavo',
    'carlos': 'Carlos',
}


def _pocket_attribution(record):
    """Same value the panel's "Atribuir a" shows; blank when unlinked."""
    return _POCKET_ATTRIBUTION_LABELS.get(record.attribution, '')


def _income_client(record):
    """Client display name; blank for the unassigned rows."""
    if not record.client_id:
        return ''
    from accounts.services.proposal_client_service import (
        build_client_display_name,
    )

    return build_client_display_name(record.client)


def _email_template_label(record):
    from content.serializers.accounting import EMAIL_TEMPLATE_LABELS

    return EMAIL_TEMPLATE_LABELS.get(record.template_key, record.template_key)


def _email_targets_label(record):
    """The records the email named, so a row stands on its own in a sheet."""
    return ' · '.join(
        target.object_repr or f'{target.entity_type} #{target.object_id}'
        for target in record.targets.all()
    )


def _change_fields_label(record):
    """Field-level diff flattened to one cell: `Campo: antes → después`."""
    return ' · '.join(
        f"{change.get('label') or change.get('field')}: "
        f"{change.get('old', '')} → {change.get('new', '')}"
        for change in (record.changes or [])
    )


EXPORT_SECTIONS = {
    'statement': {
        'title': 'Extractos TC',
        'columns': [
            ('Tarjeta', 'card_name'),
            ('Mes', 'period_date'),
            ('Estado', lambda r: r.get_status_display()),
            ('Total compras', 'purchases_total'),
            ('Pagos y abonos', 'payments_total'),
            ('Intereses y comisiones', 'interest_and_fees'),
            ('Saldo anterior', 'previous_balance'),
            ('Saldo de cierre', 'closing_balance'),
            ('Pago mínimo', 'minimum_payment'),
            ('Fecha límite', 'due_date'),
            ('Notas', 'notes'),
        ],
    },
    'statement_tx': {
        'title': 'Transacciones TC',
        'columns': [
            ('Tarjeta', lambda r: r.statement.card_name),
            ('Mes', lambda r: r.statement.period_date),
            ('Fecha', 'transaction_date'),
            ('Descripción', 'raw_description'),
            ('Comercio', 'merchant_name'),
            ('Categoría', lambda r: r.get_category_display()),
            ('Cuota', lambda r: (
                f'{r.installment_number}/{r.installments_total}'
                if r.installment_number and r.installments_total else ''
            )),
            ('Valor', 'amount'),
            ('Moneda original', 'original_currency'),
            ('Valor original', 'original_amount'),
            ('Notas', 'notes'),
        ],
    },
    'merchant_alias': {
        'title': 'Alias comercios',
        'columns': [
            ('Texto de coincidencia', 'match_text'),
            ('Comercio', 'merchant_name'),
            ('Categoría por defecto', lambda r: r.get_default_category_display()),
            ('Notas', 'notes'),
        ],
    },
    'income': {
        'title': 'Ingresos',
        'columns': [
            ('Concepto', 'concept'),
            ('Tipo', lambda r: r.get_kind_display()),
            ('Estado de cobro', _income_payment_status),
            ('Contabilidad', lambda r: r.get_ledger_display()),
            ('Mes', 'period_date'),
            ('Total', 'total_amount'),
            ('Gustavo', 'gustavo_amount'),
            ('Carlos', 'carlos_amount'),
            ('Destino', lambda r: r.get_destination_display()),
            ('Notas', 'notes'),
            # Appended, never inserted: the CSV header test pins the first
            # five columns.
            ('Cliente', _income_client),
            ('Origen', lambda r: r.get_origin_display()),
            ('Proyecto', _record_project),
            ('Período inicio', 'period_start'),
            ('Período fin', 'period_end'),
            ('Periodicidad', lambda r: (
                r.get_period_cadence_display() if r.period_cadence else ''
            )),
        ],
    },
    'expense': {
        'title': 'Gastos',
        'columns': [
            ('Concepto', 'concept'),
            ('Categoría', lambda r: r.get_category_display()),
            ('Tipo de deducción', lambda r: (
                r.get_deduction_type_display() if r.deduction_type else ''
            )),
            ('Ingreso origen', lambda r: (
                getattr(r, 'source_income_concept', None) or ''
            )),
            ('Contabilidad', lambda r: r.get_ledger_display()),
            ('Mes', 'period_date'),
            ('Total', 'total_amount'),
            ('Gustavo', 'gustavo_amount'),
            ('Carlos', 'carlos_amount'),
            ('Notas', 'notes'),
        ],
    },
    'hosting': {
        'title': 'Hostings',
        'columns': [
            ('Cliente', 'client_name'),
            ('Dominio', 'domain_url'),
            ('Valor mensual', 'monthly_value'),
            ('Modalidad', lambda r: r.payment_modality_label),
            ('Vigente desde', 'valid_from'),
            ('Vigente hasta', 'valid_to'),
            ('Ciclos', 'cycles_count'),
            ('Pago por ciclo', 'payment_per_cycle'),
            ('Total pagado', 'total_paid'),
            ('Activo', lambda r: 'Sí' if r.is_active else 'No'),
            ('Notas', 'notes'),
            # Appended, never inserted (same rule as the income section):
            # 'Cliente' above is the billing snapshot, this is the relation.
            ('Cliente vinculado', _hosting_client),
            ('Proyecto', _record_project),
        ],
    },
    'pocket': {
        'title': 'Bolsillo',
        'columns': [
            ('Concepto', 'concept'),
            ('Fecha', 'movement_date'),
            ('Tipo', lambda r: r.get_direction_display()),
            ('Valor', 'amount'),
            ('Notas', 'notes'),
            # Appended, never inserted. Both are filterable in the panel, so
            # the file has to show why the exported rows are the ones there.
            ('Atribución', _pocket_attribution),
            ('Vinculado', lambda r: 'Sí' if r.is_auto_managed else 'No'),
        ],
    },
    'recurring': {
        'title': 'Recurrentes',
        'columns': [
            ('Nombre', 'name'),
            ('Categoría', lambda r: r.category.name if r.category else ''),
            ('Precio', 'price'),
            ('Moneda', 'currency'),
            ('Precio mensual', 'monthly_price'),
            ('Equivalente COP', 'cop_equivalent'),
            ('Equiv. COP mensual', 'monthly_cop_cost'),
            ('Método de pago', lambda r: r.get_payment_method_display()),
            ('Frecuencia', lambda r: r.frequency_display),
            ('Día de cobro', 'billing_day'),
            ('Tipo de costo', lambda r: r.get_cost_type_display()),
            ('Activo', lambda r: 'Sí' if r.is_active else 'No'),
            ('Notas', 'notes'),
        ],
    },
    'ads': {
        'title': 'Ads',
        'columns': [
            ('Fecha', 'spend_date'),
            ('Plataforma', lambda r: r.get_platform_display()),
            ('Tarjeta origen', 'origin_card'),
            ('Valor', 'amount'),
            ('Notas', 'notes'),
        ],
    },
    'card_snapshot': {
        'title': 'Tarjetas',
        'columns': [
            ('Tarjeta', 'card_name'),
            ('Fecha', 'snapshot_date'),
            ('Disponible', 'available_amount'),
            ('Deuda', 'debt_amount'),
            ('Notas', 'notes'),
        ],
    },
    # The two Historial subtabs. Exported so an aviso that did go out can be
    # attached as evidence, which is the whole reason the tab is consulted.
    'email_log': {
        'title': 'Envíos',
        'columns': [
            ('Fecha', 'sent_at'),
            ('Aviso', _email_template_label),
            ('Destinatario', 'recipient'),
            ('Asunto', 'subject'),
            ('Estado', lambda r: r.get_status_display()),
            ('Error', 'error_message'),
            ('Registros', _email_targets_label),
        ],
    },
    'change_log': {
        'title': 'Cambios',
        'columns': [
            ('Fecha', 'created_at'),
            ('Usuario', 'actor_username'),
            ('Entidad', lambda r: r.get_entity_type_display()),
            ('Registro', 'object_repr'),
            ('Acción', lambda r: r.get_action_display()),
            ('Campos', _change_fields_label),
        ],
    },
}


def _cell(record, accessor):
    if callable(accessor):
        return accessor(record)
    return getattr(record, accessor)


def iter_rows(section_key, queryset):
    """Yield one list of cell values per record."""
    columns = EXPORT_SECTIONS[section_key]['columns']
    for record in queryset:
        yield [_cell(record, accessor) for _header, accessor in columns]


def section_headers(section_key):
    return [header for header, _accessor in EXPORT_SECTIONS[section_key]['columns']]


def write_csv(response, section_key, queryset):
    """Write the section as CSV into an HttpResponse (BOM for Excel)."""
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow(section_headers(section_key))
    for row in iter_rows(section_key, queryset):
        writer.writerow(row)
    return response


def _write_sheet(sheet, section_key, queryset):
    headers = section_headers(section_key)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = HEADER_FONT
    money_columns = set()
    for row in iter_rows(section_key, queryset):
        sheet.append(row)
        for index, value in enumerate(row, start=1):
            if isinstance(value, Decimal):
                money_columns.add(index)
    for index in money_columns:
        for cell in sheet.iter_cols(
            min_col=index, max_col=index, min_row=2,
        ):
            for c in cell:
                c.number_format = MONEY_FORMAT
    for index, header in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = max(
            14, len(header) + 4,
        )


def write_xlsx(section_key, queryset):
    """Return a single-sheet Workbook for the section."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = EXPORT_SECTIONS[section_key]['title']
    _write_sheet(sheet, section_key, queryset)
    return workbook


def _append_summary_block(sheet, title, rows):
    sheet.append([title])
    sheet[sheet.max_row][0].font = HEADER_FONT
    for label, value in rows:
        sheet.append([label, value])
        if isinstance(value, Decimal):
            sheet.cell(row=sheet.max_row, column=2).number_format = MONEY_FORMAT
    sheet.append([])


def build_workbook(year, sections_querysets):
    """Full accounting workbook: summary sheet + one sheet per section.

    `sections_querysets` maps section keys to their (already filtered)
    querysets so the view controls year scoping per section.
    """
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = 'Resumen'
    summary = dashboard_summary(year)

    _append_summary_block(summary_sheet, f'Resumen {year}', [
        ('Ingresos esperados', summary['expected_total']),
        ('Ingresos líquidos', summary['liquid_total']),
        ('Gastos', summary['expenses_total']),
        ('Utilidad esperada', summary['expected_utility']),
        ('Utilidad líquida', summary['liquid_utility']),
        ('Bolsillo ProjectApp', summary['pocket_balance']),
        ('Costo operativo mensual', summary['recurring_monthly_cost']),
    ])

    partner_labels = {
        'gustavo': 'Gustavo',
        'carlos': 'Carlos',
        'company': 'ProjectApp (Empresa)',
    }
    for key, label in partner_labels.items():
        partner = summary['partners'][key]
        _append_summary_block(summary_sheet, label, [
            ('Esperado', partner['expected']),
            ('Líquido', partner['liquid']),
            ('Gastos', partner['expenses']),
            ('Neto', partner['net']),
        ])

    summary_sheet.append(['Detalle mensual'])
    summary_sheet[summary_sheet.max_row][0].font = HEADER_FONT
    summary_sheet.append(['Mes', 'Esperado', 'Líquido', 'Gastos', 'Utilidad'])
    for cell in summary_sheet[summary_sheet.max_row]:
        cell.font = HEADER_FONT
    for month in summary['monthly']:
        summary_sheet.append([
            month['label'], month['expected'], month['liquid'],
            month['expenses'], month['utility'],
        ])
        for column in range(2, 6):
            summary_sheet.cell(
                row=summary_sheet.max_row, column=column,
            ).number_format = MONEY_FORMAT
    summary_sheet.append([])

    if summary['latest_card_snapshots']:
        summary_sheet.append(['Tarjetas (último registro)'])
        summary_sheet[summary_sheet.max_row][0].font = HEADER_FONT
        summary_sheet.append(['Tarjeta', 'Fecha', 'Disponible', 'Deuda'])
        for cell in summary_sheet[summary_sheet.max_row]:
            cell.font = HEADER_FONT
        for snapshot in summary['latest_card_snapshots']:
            summary_sheet.append([
                snapshot['card_name'], snapshot['snapshot_date'],
                snapshot['available_amount'], snapshot['debt_amount'],
            ])
            for column in (3, 4):
                summary_sheet.cell(
                    row=summary_sheet.max_row, column=column,
                ).number_format = MONEY_FORMAT

    summary_sheet.column_dimensions['A'].width = 28
    for letter in ('B', 'C', 'D', 'E'):
        summary_sheet.column_dimensions[letter].width = 16

    for section_key, queryset in sections_querysets.items():
        sheet = workbook.create_sheet(EXPORT_SECTIONS[section_key]['title'])
        _write_sheet(sheet, section_key, queryset)

    return workbook


def export_filename(stem, extension):
    return f'{stem}_{today_bogota().strftime("%Y%m%d")}.{extension}'
